"""
Financial Ratio Analysis Tool — for student learning
------------------------------------------------------
Students manually enter balance sheet & income statement figures for
TWO years side by side. The app computes 18 standard ratios across
four categories, compares Year 1 vs Year 2, interprets each ratio in
plain language, and lets students download the comparison as CSV,
Excel (.xlsx), or a formatted PDF report.

Run with:  streamlit run ratio_analysis_app.py
"""

import io

import pandas as pd
import streamlit as st

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter as PAGE_SIZE
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

# ============================================================
# PAGE CONFIG
# ============================================================
st.set_page_config(
    page_title="Financial Ratio Analysis Tool",
    page_icon="📊",
    layout="wide",
)

# ============================================================
# GENERIC HELPERS
# ============================================================

def safe_div(numerator, denominator):
    """Divide safely; return None if denominator is zero/blank."""
    if denominator in (0, None):
        return None
    return numerator / denominator


def fmt(value, suffix=""):
    """Format a ratio value for display, handling None gracefully."""
    if value is None:
        return "N/A"
    if suffix == "%":
        return f"{value * 100:,.1f}%"
    if suffix == " days":
        return f"{value:,.0f} days"
    if suffix == "x":
        return f"{value:,.2f}x"
    return f"{value:,.2f}"


def fmt_delta(delta, suffix):
    """Format a year-over-year change for st.metric's delta display."""
    if delta is None:
        return None
    if suffix == "%":
        return f"{delta * 100:+.1f} pp"  # percentage-POINT change
    if suffix == " days":
        return f"{delta:+.0f} days"
    if suffix == "x":
        return f"{delta:+.2f}x"
    return f"{delta:+.2f}"


NA_MSG = "Can't be calculated — check that the denominator in the formula above isn't zero or blank."

# ============================================================
# INTERPRETATION FUNCTIONS
# Each returns (level, message). level in {good, ok, warning, na}
# Interpretations are applied to the *current* (Year 2) value.
# Thresholds are common textbook rules of thumb — real benchmarks
# vary by industry, so treat these as a starting point for discussion.
# ============================================================

def interp_current_ratio(v):
    if v is None:
        return "na", NA_MSG
    if v >= 2:
        return "good", "Comfortably above 2, meaning current assets cover current liabilities more than twice over. Strong short-term liquidity, though very high values can also mean idle cash or excess inventory."
    if v >= 1:
        return "ok", "Between 1 and 2 — current assets cover current liabilities, but the safety margin is moderate. Worth comparing against industry peers."
    return "warning", "Below 1 — current liabilities exceed current assets. This is a liquidity warning sign; the company may struggle to meet short-term obligations."


def interp_quick_ratio(v):
    if v is None:
        return "na", NA_MSG
    if v >= 1:
        return "good", "At or above 1 — the company can cover current liabilities without relying on selling inventory. Solid immediate liquidity."
    if v >= 0.7:
        return "ok", "Slightly below 1 — liquid assets almost cover current liabilities. Reasonable, but leaves little room for error."
    return "warning", "Well below 1 — the company depends heavily on inventory or other slow-moving assets to meet short-term debts."


def interp_cash_ratio(v):
    if v is None:
        return "na", NA_MSG
    if v >= 0.5:
        return "good", "High cash coverage of current liabilities — very conservative liquidity position. Could also suggest cash is not being put to productive use."
    if v >= 0.2:
        return "ok", "Moderate cash coverage — a reasonable cash buffer against current liabilities."
    return "warning", "Low cash coverage — the company holds little cash relative to what it owes in the short term."


def interp_gross_margin(v):
    if v is None:
        return "na", NA_MSG
    if v >= 0.40:
        return "good", "A high gross margin suggests strong pricing power or low production costs relative to sales."
    if v >= 0.20:
        return "ok", "A moderate gross margin — typical of many manufacturing and retail businesses."
    return "warning", "A low gross margin leaves little cushion after production costs — worth checking cost control and pricing strategy."


def interp_operating_margin(v):
    if v is None:
        return "na", NA_MSG
    if v >= 0.15:
        return "good", "Strong operating margin — core business operations are generating healthy profit before interest and tax."
    if v >= 0.05:
        return "ok", "Positive but modest operating margin — operating expenses are consuming a large share of gross profit."
    return "warning", "Weak or negative operating margin — operating costs are eating heavily into profitability."


def interp_net_margin(v):
    if v is None:
        return "na", NA_MSG
    if v >= 0.10:
        return "good", "A healthy net margin — for every dollar of sales, a solid portion becomes bottom-line profit."
    if v >= 0.02:
        return "ok", "A thin but positive net margin — profitable, but with limited buffer against rising costs."
    return "warning", "Very low or negative net margin — the company is barely profitable (or is losing money) after all expenses."


def interp_roa(v):
    if v is None:
        return "na", NA_MSG
    if v >= 0.10:
        return "good", "The company is generating strong profit relative to the assets it owns — efficient use of the asset base."
    if v >= 0.03:
        return "ok", "Modest returns on assets — assets are generating some profit, but there may be room to use them more efficiently."
    return "warning", "Low or negative ROA — assets are not generating much profit relative to their size."


def interp_roe(v):
    if v is None:
        return "na", NA_MSG
    if v >= 0.15:
        return "good", "Strong return for shareholders — the company is generating healthy profit relative to equity invested."
    if v >= 0.05:
        return "ok", "Positive but modest return to shareholders — acceptable, but compare against the return investors could get elsewhere."
    return "warning", "Low or negative ROE — shareholders are earning little (or losing value) on their investment in the company."


def interp_asset_turnover(v):
    if v is None:
        return "na", NA_MSG
    if v >= 1.0:
        return "good", "The company generates at least a dollar of sales for every dollar of assets — efficient asset use."
    if v >= 0.5:
        return "ok", "Moderate efficiency in using assets to generate sales. Typical for capital-intensive industries."
    return "warning", "Low asset turnover — the company is generating relatively little sales from its asset base, which may indicate underused or excess assets."


def interp_inventory_turnover(v):
    if v is None:
        return "na", NA_MSG
    if v >= 8:
        return "good", "Inventory is selling and being replaced quickly — efficient inventory management."
    if v >= 3:
        return "ok", "Reasonable inventory turnover — goods move at a moderate pace."
    return "warning", "Low inventory turnover — inventory is sitting for a long time before being sold, tying up cash and raising obsolescence risk."


def interp_dio(v):
    if v is None:
        return "na", NA_MSG
    if v <= 45:
        return "good", "Inventory is held for a relatively short period before sale — efficient inventory cycle."
    if v <= 90:
        return "ok", "A moderate number of days to sell through inventory."
    return "warning", "Inventory sits for a long time before being sold — this ties up working capital."


def interp_receivables_turnover(v):
    if v is None:
        return "na", NA_MSG
    if v >= 10:
        return "good", "Receivables are collected quickly — effective credit and collections management."
    if v >= 4:
        return "ok", "Moderate collection speed on credit sales."
    return "warning", "Receivables turn over slowly — customers are taking a long time to pay, which can strain cash flow."


def interp_dso(v):
    if v is None:
        return "na", NA_MSG
    if v <= 45:
        return "good", "Customers pay relatively quickly after a credit sale — healthy collections."
    if v <= 90:
        return "ok", "A moderate collection period — worth monitoring against the company's credit terms."
    return "warning", "A long collection period — cash is tied up in receivables for a long time, which can create cash flow strain."


def interp_debt_ratio(v):
    if v is None:
        return "na", NA_MSG
    if v <= 0.4:
        return "good", "A relatively low proportion of assets is financed by debt — conservative financing and lower financial risk."
    if v <= 0.6:
        return "ok", "A moderate reliance on debt financing — common for many established companies."
    return "warning", "A high proportion of assets is financed by debt — this raises financial risk and sensitivity to interest rate or earnings shocks."


def interp_debt_to_equity(v):
    if v is None:
        return "na", NA_MSG
    if v <= 1.0:
        return "good", "Debt is less than or equal to equity — a relatively conservative capital structure."
    if v <= 2.0:
        return "ok", "Debt moderately exceeds equity — leverage is elevated but not extreme; depends heavily on the industry norm."
    return "warning", "Debt significantly exceeds equity — high financial leverage increases risk to shareholders and creditors alike."


def interp_equity_multiplier(v):
    if v is None:
        return "na", NA_MSG
    if v <= 2.0:
        return "good", "Assets are financed mostly by equity rather than debt — lower financial leverage."
    if v <= 3.0:
        return "ok", "A moderate level of financial leverage — a meaningful portion of assets is debt-financed."
    return "warning", "High financial leverage — a large share of assets is financed by debt relative to equity, amplifying both potential returns and risk."


def interp_interest_coverage(v):
    if v is None:
        return "na", NA_MSG
    if v >= 5:
        return "good", "Operating profit covers interest expense many times over — little risk of default on interest payments."
    if v >= 1.5:
        return "ok", "Operating profit covers interest, but the cushion is moderate — a downturn in earnings could pressure debt payments."
    return "warning", "Operating profit barely covers (or doesn't cover) interest expense — a serious warning sign for solvency."


LEVEL_LABEL = {"good": "✅ Strong", "ok": "🟡 Moderate", "warning": "🔴 Weak", "na": "⚪ N/A"}
# Plain-text versions for the PDF export: ReportLab's built-in fonts can't render
# emoji or Unicode arrow/triangle glyphs, so the PDF uses these instead.
LEVEL_LABEL_PLAIN = {"good": "Strong", "ok": "Moderate", "warning": "Weak", "na": "N/A"}
TREND_LABEL = {"improved": "▲ Improved", "declined": "▼ Declined", "flat": "→ Flat", "na": "N/A"}
TREND_LABEL_PLAIN = {"improved": "Improved", "declined": "Declined", "flat": "Flat", "na": "N/A"}

# ============================================================
# RATIO DEFINITIONS (single source of truth used everywhere below)
# ============================================================
RATIO_DEFS = [
    dict(key="current_ratio", category="Liquidity", name="Current Ratio",
         formula="Current Assets ÷ Current Liabilities", suffix="", lower_better=False, interp=interp_current_ratio),
    dict(key="quick_ratio", category="Liquidity", name="Quick Ratio (Acid-Test)",
         formula="(Current Assets − Inventory) ÷ Current Liabilities", suffix="", lower_better=False, interp=interp_quick_ratio),
    dict(key="cash_ratio", category="Liquidity", name="Cash Ratio",
         formula="Cash ÷ Current Liabilities", suffix="", lower_better=False, interp=interp_cash_ratio),

    dict(key="gross_margin", category="Profitability", name="Gross Profit Margin",
         formula="Gross Profit ÷ Sales", suffix="%", lower_better=False, interp=interp_gross_margin),
    dict(key="operating_margin", category="Profitability", name="Operating Profit Margin",
         formula="EBIT ÷ Sales", suffix="%", lower_better=False, interp=interp_operating_margin),
    dict(key="net_margin", category="Profitability", name="Net Profit Margin",
         formula="Net Income ÷ Sales", suffix="%", lower_better=False, interp=interp_net_margin),
    dict(key="roa", category="Profitability", name="Return on Assets (ROA)",
         formula="Net Income ÷ Total Assets", suffix="%", lower_better=False, interp=interp_roa),
    dict(key="roe", category="Profitability", name="Return on Equity (ROE)",
         formula="Net Income ÷ Shareholders' Equity", suffix="%", lower_better=False, interp=interp_roe),

    dict(key="asset_turnover", category="Efficiency", name="Asset Turnover",
         formula="Sales ÷ Total Assets", suffix="x", lower_better=False, interp=interp_asset_turnover),
    dict(key="inventory_turnover", category="Efficiency", name="Inventory Turnover",
         formula="COGS ÷ Inventory", suffix="x", lower_better=False, interp=interp_inventory_turnover),
    dict(key="dio", category="Efficiency", name="Days Inventory Outstanding",
         formula="365 ÷ Inventory Turnover", suffix=" days", lower_better=True, interp=interp_dio),
    dict(key="receivables_turnover", category="Efficiency", name="Receivables Turnover",
         formula="Sales ÷ Accounts Receivable", suffix="x", lower_better=False, interp=interp_receivables_turnover),
    dict(key="dso", category="Efficiency", name="Days Sales Outstanding (DSO)",
         formula="365 ÷ Receivables Turnover", suffix=" days", lower_better=True, interp=interp_dso),

    dict(key="debt_ratio", category="Solvency", name="Debt Ratio",
         formula="Total Liabilities ÷ Total Assets", suffix="%", lower_better=True, interp=interp_debt_ratio),
    dict(key="debt_to_equity", category="Solvency", name="Debt-to-Equity Ratio",
         formula="Total Liabilities ÷ Equity", suffix="x", lower_better=True, interp=interp_debt_to_equity),
    dict(key="equity_multiplier", category="Solvency", name="Equity Multiplier",
         formula="Total Assets ÷ Equity", suffix="x", lower_better=True, interp=interp_equity_multiplier),
    dict(key="interest_coverage", category="Solvency", name="Interest Coverage Ratio",
         formula="EBIT ÷ Interest Expense", suffix="x", lower_better=False, interp=interp_interest_coverage),
]
CATEGORIES = ["Liquidity", "Profitability", "Efficiency", "Solvency"]

# ============================================================
# DERIVED FIGURES & RATIO CALCULATION
# ============================================================

def compute_derived(v):
    """Given raw line items for one year, compute totals & income lines."""
    tca = v["cash"] + v["receivables"] + v["inventory"] + v["other_ca"]
    ta = tca + v["fixed_assets"]
    tcl = v["payables"] + v["other_cl"]
    tl = tcl + v["long_term_debt"]
    gross_profit = v["sales"] - v["cogs"]
    ebit = gross_profit - v["operating_expenses"]
    ebt = ebit - v["interest_expense"]
    net_income = ebt - v["tax_expense"]
    out = dict(v)
    out.update(
        total_current_assets=tca, total_assets=ta,
        total_current_liabilities=tcl, total_liabilities=tl,
        gross_profit=gross_profit, ebit=ebit, ebt=ebt, net_income=net_income,
    )
    return out


def compute_ratios(d):
    r = {}
    r["current_ratio"] = safe_div(d["total_current_assets"], d["total_current_liabilities"])
    r["quick_ratio"] = safe_div(d["total_current_assets"] - d["inventory"], d["total_current_liabilities"])
    r["cash_ratio"] = safe_div(d["cash"], d["total_current_liabilities"])

    r["gross_margin"] = safe_div(d["gross_profit"], d["sales"])
    r["operating_margin"] = safe_div(d["ebit"], d["sales"])
    r["net_margin"] = safe_div(d["net_income"], d["sales"])
    r["roa"] = safe_div(d["net_income"], d["total_assets"])
    r["roe"] = safe_div(d["net_income"], d["equity"])

    r["asset_turnover"] = safe_div(d["sales"], d["total_assets"])
    r["inventory_turnover"] = safe_div(d["cogs"], d["inventory"])
    r["dio"] = safe_div(365, r["inventory_turnover"]) if r["inventory_turnover"] else None
    r["receivables_turnover"] = safe_div(d["sales"], d["receivables"])
    r["dso"] = safe_div(365, r["receivables_turnover"]) if r["receivables_turnover"] else None

    r["debt_ratio"] = safe_div(d["total_liabilities"], d["total_assets"])
    r["debt_to_equity"] = safe_div(d["total_liabilities"], d["equity"])
    r["equity_multiplier"] = safe_div(d["total_assets"], d["equity"])
    r["interest_coverage"] = safe_div(d["ebit"], d["interest_expense"])
    return r


# ============================================================
# EXCEL EXPORT
# ============================================================

def build_excel_with_raw(company_name, y1_label, y2_label, export_rows, raw_lines):
    """Build the workbook with both the ratio comparison and a raw-data sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ratio Comparison"

    title_font = Font(name="Arial", size=14, bold=True)
    subtitle_font = Font(name="Arial", size=10, italic=True, color="666666")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    normal_font = Font(name="Arial", size=10)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:G1")
    ws["A1"] = f"{company_name} — Financial Ratio Analysis"
    ws["A1"].font = title_font
    ws.merge_cells("A2:G2")
    ws["A2"] = f"{y1_label} vs {y2_label}"
    ws["A2"].font = subtitle_font

    headers = ["Category", "Ratio", y1_label, y2_label, "Change", "Trend", "Assessment (latest year)"]
    header_row = 4
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    number_formats = {"": "0.00", "%": "0.0%", "x": '0.00"x"', " days": '0" days"'}
    r = header_row
    for row in export_rows:
        r += 1
        nf = number_formats.get(row["suffix"], "0.00")
        ws.cell(row=r, column=1, value=row["category"]).font = normal_font
        ws.cell(row=r, column=2, value=row["ratio"]).font = normal_font
        for col_idx, val in ((3, row["y1"]), (4, row["y2"]), (5, row["change"])):
            cell = ws.cell(row=r, column=col_idx, value=("N/A" if val is None else val))
            cell.font = normal_font
            if val is not None:
                cell.number_format = nf
            cell.alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=6, value=row["trend"]).font = normal_font
        ws.cell(row=r, column=7, value=row["assessment"]).font = normal_font
        for col_idx in range(1, 8):
            ws.cell(row=r, column=col_idx).border = border

    widths = [14, 30, 14, 14, 12, 14, 24]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"

    # ---- Raw data sheet ----
    ws2 = wb.create_sheet("Raw Data")
    ws2.merge_cells("A1:C1")
    ws2["A1"] = f"{company_name} — Raw Inputs"
    ws2["A1"].font = title_font
    raw_headers = ["Line Item", y1_label, y2_label]
    for ci, h in enumerate(raw_headers, start=1):
        cell = ws2.cell(row=3, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    rr = 3
    for label, y1v, y2v in raw_lines:
        rr += 1
        ws2.cell(row=rr, column=1, value=label).font = normal_font
        c2 = ws2.cell(row=rr, column=2, value=y1v)
        c2.font = normal_font
        c2.number_format = "#,##0.00"
        c3 = ws2.cell(row=rr, column=3, value=y2v)
        c3.font = normal_font
        c3.number_format = "#,##0.00"
        for ci in range(1, 4):
            ws2.cell(row=rr, column=ci).border = border
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ============================================================
# PDF EXPORT
# ============================================================

def build_pdf_report(company_name, y1_label, y2_label, export_rows):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=PAGE_SIZE,
        leftMargin=36, rightMargin=36, topMargin=48, bottomMargin=36,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("ReportTitle", parent=styles["Title"], fontSize=16, spaceAfter=4)
    sub_style = ParagraphStyle("ReportSub", parent=styles["Normal"], fontSize=9, textColor=colors.grey, spaceAfter=14)
    section_style = ParagraphStyle(
        "Section", parent=styles["Heading2"], fontSize=12,
        spaceBefore=14, spaceAfter=6, textColor=colors.HexColor("#1F4E78"),
    )
    footnote_style = ParagraphStyle("Footnote", parent=styles["Normal"], fontSize=7.5, textColor=colors.grey)

    story = [
        Paragraph(company_name, title_style),
        Paragraph(f"Financial Ratio Analysis &mdash; {y1_label} vs {y2_label}", sub_style),
    ]

    for cat in CATEGORIES:
        rows = [r for r in export_rows if r["category"] == cat]
        if not rows:
            continue
        story.append(Paragraph(cat, section_style))
        data = [["Ratio", y1_label, y2_label, "Change", "Trend", "Assessment"]]
        for r in rows:
            if r["change"] is None:
                change_str = "N/A"
            elif r["suffix"] == "%":
                change_str = f"{r['change'] * 100:+.1f} pp"
            else:
                change_str = fmt(r["change"], r["suffix"])
            data.append([
                r["ratio"],
                fmt(r["y1"], r["suffix"]),
                fmt(r["y2"], r["suffix"]),
                change_str,
                TREND_LABEL_PLAIN[r["trend_code"]],
                LEVEL_LABEL_PLAIN[r["level"]],
            ])
        table = Table(data, colWidths=[148, 62, 62, 58, 68, 78], repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 8.5),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F6FA")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(table)

    story.append(Spacer(1, 14))
    story.append(Paragraph(
        "Assessment thresholds are general classroom rules of thumb; real-world benchmarks vary by industry. "
        "Generated by the student Financial Ratio Analysis Tool.",
        footnote_style,
    ))

    doc.build(story)
    buf.seek(0)
    return buf


# ============================================================
# RENDERING HELPERS
# ============================================================

def render_ratio(col, rdef, val_y1, val_y2, y1_label, y2_label):
    with col:
        level, message = rdef["interp"](val_y2)
        delta_val = (val_y2 - val_y1) if (val_y1 is not None and val_y2 is not None) else None
        st.metric(
            rdef["name"],
            fmt(val_y2, rdef["suffix"]),
            delta=fmt_delta(delta_val, rdef["suffix"]),
            delta_color=("inverse" if rdef["lower_better"] else "normal"),
        )
        st.caption(f"Formula: {rdef['formula']}")
        st.caption(f"{y1_label}: {fmt(val_y1, rdef['suffix'])}   →   {y2_label}: {fmt(val_y2, rdef['suffix'])}")
        box = {"good": st.success, "ok": st.info, "warning": st.warning, "na": st.error}[level]
        box(message)
        st.write("")


def render_category_tab(category, ratios_y1, ratios_y2, y1_label, y2_label, col_count=3):
    defs = [r for r in RATIO_DEFS if r["category"] == category]
    for i in range(0, len(defs), col_count):
        chunk = defs[i:i + col_count]
        cols = st.columns(len(chunk))
        for c, rdef in zip(cols, chunk):
            render_ratio(c, rdef, ratios_y1[rdef["key"]], ratios_y2[rdef["key"]], y1_label, y2_label)


# ============================================================
# SIDEBAR
# ============================================================
with st.sidebar:
    st.header("📘 How to use this tool")
    st.markdown(
        """
        1. Go to **Enter Data** and fill in the two spreadsheet-style
           tables — one column per year — using figures from a
           balance sheet and income statement (sample values are
           pre-filled so you can explore first).
        2. Browse the **Liquidity**, **Profitability**, **Efficiency**,
           and **Solvency** tabs to see each ratio for both years side
           by side, with its formula, year-over-year change, and a
           plain-language interpretation.
        3. Check the **Summary Dashboard** for a one-page comparison,
           and download it as **CSV**, **Excel**, or a formatted
           **PDF report**.

        *Note:* Interpretation thresholds are general classroom
        rules of thumb. Real-world benchmarks vary a lot by industry —
        use these as a starting point for discussion, not a strict rule.
        """
    )
    company_name = st.text_input("Company name (optional)", value="Sample Company Inc.")
    year1_label = st.text_input("Year 1 label", value="FY2024")
    year2_label = st.text_input("Year 2 label", value="FY2025")

st.title("📊 Financial Ratio Analysis Tool")
st.caption("Enter two years of data, then compare ratios and see what each one means.")

tab_input, tab_liq, tab_prof, tab_eff, tab_solv, tab_summary = st.tabs(
    ["📥 Enter Data", "💧 Liquidity", "💰 Profitability", "⚙️ Efficiency", "🏦 Solvency", "📋 Summary Dashboard"]
)

# ============================================================
# TAB: ENTER DATA
# ============================================================
BS_ITEMS = [
    ("Cash & Cash Equivalents", "cash", 15000.0, 18000.0),
    ("Accounts Receivable", "receivables", 25000.0, 28000.0),
    ("Inventory", "inventory", 30000.0, 32000.0),
    ("Other Current Assets", "other_ca", 5000.0, 5000.0),
    ("Net Fixed / Non-Current Assets", "fixed_assets", 120000.0, 125000.0),
    ("Accounts Payable", "payables", 18000.0, 19000.0),
    ("Other Current Liabilities (e.g. short-term notes)", "other_cl", 7000.0, 7500.0),
    ("Long-Term Debt", "long_term_debt", 50000.0, 45000.0),
    ("Shareholders' Equity", "equity", 120000.0, 136500.0),
]
IS_ITEMS = [
    ("Net Sales / Revenue", "sales", 250000.0, 280000.0),
    ("Cost of Goods Sold (COGS)", "cogs", 150000.0, 163000.0),
    ("Operating Expenses (SG&A, excl. interest & tax)", "operating_expenses", 60000.0, 65000.0),
    ("Interest Expense", "interest_expense", 5000.0, 4500.0),
    ("Tax Expense", "tax_expense", 8000.0, 10000.0),
]

with tab_input:
    st.subheader("Balance Sheet & Income Statement — Two-Year Entry")
    st.write(
        f"Enter figures for **{year1_label}** and **{year2_label}** directly in the tables "
        "below (click a cell to edit, like a mini spreadsheet). Totals, gross profit, EBIT, "
        "and net income are calculated automatically for each year."
    )

    col_bs, col_is = st.columns(2)

    with col_bs:
        st.markdown("### 🏦 Balance Sheet")
        default_bs_df = pd.DataFrame({
            "Item": [x[0] for x in BS_ITEMS],
            "Year 1": [x[2] for x in BS_ITEMS],
            "Year 2": [x[3] for x in BS_ITEMS],
        })
        edited_bs = st.data_editor(
            default_bs_df,
            key="bs_editor",
            hide_index=True,
            num_rows="fixed",
            use_container_width=True,
            column_config={
                "Item": st.column_config.TextColumn("Line Item", disabled=True),
                "Year 1": st.column_config.NumberColumn(year1_label, format="%.2f"),
                "Year 2": st.column_config.NumberColumn(year2_label, format="%.2f"),
            },
        )

    with col_is:
        st.markdown("### 💵 Income Statement")
        default_is_df = pd.DataFrame({
            "Item": [x[0] for x in IS_ITEMS],
            "Year 1": [x[2] for x in IS_ITEMS],
            "Year 2": [x[3] for x in IS_ITEMS],
        })
        edited_is = st.data_editor(
            default_is_df,
            key="is_editor",
            hide_index=True,
            num_rows="fixed",
            use_container_width=True,
            column_config={
                "Item": st.column_config.TextColumn("Line Item", disabled=True),
                "Year 1": st.column_config.NumberColumn(year1_label, format="%.2f"),
                "Year 2": st.column_config.NumberColumn(year2_label, format="%.2f"),
            },
        )
        st.caption(
            "Gross Profit, Operating Income (EBIT), and Net Income are calculated "
            "automatically:\n\n"
            "Gross Profit = Sales − COGS\n\n"
            "EBIT = Gross Profit − Operating Expenses\n\n"
            "Net Income = EBIT − Interest − Tax"
        )

    bs_keys = [x[1] for x in BS_ITEMS]
    is_keys = [x[1] for x in IS_ITEMS]
    values_y1 = {k: float(v) for k, v in zip(bs_keys, edited_bs["Year 1"])}
    values_y1.update({k: float(v) for k, v in zip(is_keys, edited_is["Year 1"])})
    values_y2 = {k: float(v) for k, v in zip(bs_keys, edited_bs["Year 2"])}
    values_y2.update({k: float(v) for k, v in zip(is_keys, edited_is["Year 2"])})

    d1 = compute_derived(values_y1)
    d2 = compute_derived(values_y2)

    st.markdown("### Quick Check")
    qc1, qc2 = st.columns(2)
    for col, d, label in ((qc1, d1, year1_label), (qc2, d2, year2_label)):
        with col:
            st.markdown(f"**{label}**")
            m1, m2 = st.columns(2)
            m1.metric("Total Current Assets", f"{d['total_current_assets']:,.0f}")
            m1.metric("Total Assets", f"{d['total_assets']:,.0f}")
            m2.metric("Total Current Liabilities", f"{d['total_current_liabilities']:,.0f}")
            m2.metric("Total Liabilities", f"{d['total_liabilities']:,.0f}")
            st.metric("Net Income", f"{d['net_income']:,.0f}")
            balance_check = d["total_assets"] - (d["total_liabilities"] + d["equity"])
            if abs(balance_check) < 0.01:
                st.success("✅ Balances: Assets = Liabilities + Equity")
            else:
                st.warning(f"⚠️ Off by {balance_check:,.2f} — adjust Equity so the books balance.")

# Recompute outside the tab so all other tabs can use it regardless of tab render order
bs_keys = [x[1] for x in BS_ITEMS]
is_keys = [x[1] for x in IS_ITEMS]
values_y1 = {k: float(v) for k, v in zip(bs_keys, edited_bs["Year 1"])}
values_y1.update({k: float(v) for k, v in zip(is_keys, edited_is["Year 1"])})
values_y2 = {k: float(v) for k, v in zip(bs_keys, edited_bs["Year 2"])}
values_y2.update({k: float(v) for k, v in zip(is_keys, edited_is["Year 2"])})

d1 = compute_derived(values_y1)
d2 = compute_derived(values_y2)
ratios_y1 = compute_ratios(d1)
ratios_y2 = compute_ratios(d2)

# ============================================================
# TAB: LIQUIDITY
# ============================================================
with tab_liq:
    st.subheader("💧 Liquidity Ratios")
    st.write("Liquidity ratios measure whether a company can meet its short-term obligations.")
    render_category_tab("Liquidity", ratios_y1, ratios_y2, year1_label, year2_label)

    nwc1 = d1["total_current_assets"] - d1["total_current_liabilities"]
    nwc2 = d2["total_current_assets"] - d2["total_current_liabilities"]
    st.info(
        f"**Net Working Capital** = Current Assets − Current Liabilities  \n"
        f"{year1_label}: **{nwc1:,.0f}**  →  {year2_label}: **{nwc2:,.0f}**. "
        f"{'Positive working capital means short-term assets exceed short-term obligations.' if nwc2 >= 0 else 'Negative working capital means short-term obligations exceed short-term assets — a liquidity concern.'}"
    )

# ============================================================
# TAB: PROFITABILITY
# ============================================================
with tab_prof:
    st.subheader("💰 Profitability Ratios")
    st.write("Profitability ratios measure how well a company converts sales and assets into profit.")
    render_category_tab("Profitability", ratios_y1, ratios_y2, year1_label, year2_label)

# ============================================================
# TAB: EFFICIENCY
# ============================================================
with tab_eff:
    st.subheader("⚙️ Efficiency (Activity) Ratios")
    st.write("Efficiency ratios measure how well a company uses its assets to generate sales.")
    render_category_tab("Efficiency", ratios_y1, ratios_y2, year1_label, year2_label)

# ============================================================
# TAB: SOLVENCY
# ============================================================
with tab_solv:
    st.subheader("🏦 Solvency (Leverage) Ratios")
    st.write("Solvency ratios measure long-term financial risk and reliance on debt financing.")
    render_category_tab("Solvency", ratios_y1, ratios_y2, year1_label, year2_label)

# ============================================================
# TAB: SUMMARY DASHBOARD
# ============================================================
with tab_summary:
    st.subheader("📋 Summary Dashboard")
    st.caption(f"{company_name} — {year1_label} vs {year2_label}")

    export_rows = []
    for rdef in RATIO_DEFS:
        v1 = ratios_y1[rdef["key"]]
        v2 = ratios_y2[rdef["key"]]
        level, _ = rdef["interp"](v2)
        change = (v2 - v1) if (v1 is not None and v2 is not None) else None
        if change is None:
            trend_code = "na"
        elif abs(change) < 1e-9:
            trend_code = "flat"
        else:
            improved = (change < 0) if rdef["lower_better"] else (change > 0)
            trend_code = "improved" if improved else "declined"
        export_rows.append(dict(
            category=rdef["category"], ratio=rdef["name"], suffix=rdef["suffix"],
            y1=v1, y2=v2, change=change,
            level=level, trend_code=trend_code,
            trend=TREND_LABEL[trend_code], assessment=LEVEL_LABEL[level],
        ))

    display_rows = [
        {
            "Category": r["category"],
            "Ratio": r["ratio"],
            year1_label: fmt(r["y1"], r["suffix"]),
            year2_label: fmt(r["y2"], r["suffix"]),
            "Change": ("N/A" if r["change"] is None else (
                f"{r['change'] * 100:+.1f} pp" if r["suffix"] == "%" else fmt(r["change"], r["suffix"])
            )),
            "Trend": r["trend"],
            "Assessment": r["assessment"],
        }
        for r in export_rows
    ]
    df_display = pd.DataFrame(display_rows)
    st.dataframe(df_display, use_container_width=True, hide_index=True)

    level_counts = pd.Series([r["assessment"] for r in export_rows]).value_counts()
    cc1, cc2, cc3, cc4 = st.columns(4)
    cc1.metric("✅ Strong", int(level_counts.get("✅ Strong", 0)))
    cc2.metric("🟡 Moderate", int(level_counts.get("🟡 Moderate", 0)))
    cc3.metric("🔴 Weak", int(level_counts.get("🔴 Weak", 0)))
    cc4.metric("⚪ Not calculable", int(level_counts.get("⚪ N/A", 0)))

    st.markdown("#### ⬇️ Download this comparison")
    dl1, dl2, dl3 = st.columns(3)

    with dl1:
        csv_bytes = df_display.to_csv(index=False).encode("utf-8")
        st.download_button(
            "CSV", data=csv_bytes,
            file_name=f"{company_name.replace(' ', '_')}_ratio_comparison.csv",
            mime="text/csv", use_container_width=True,
        )

    with dl2:
        raw_lines = []
        for label, key, _, _ in BS_ITEMS + IS_ITEMS:
            raw_lines.append((label, values_y1[key], values_y2[key]))
        excel_buf = build_excel_with_raw(company_name, year1_label, year2_label, export_rows, raw_lines)
        st.download_button(
            "Excel (.xlsx)", data=excel_buf,
            file_name=f"{company_name.replace(' ', '_')}_ratio_comparison.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    with dl3:
        pdf_buf = build_pdf_report(company_name, year1_label, year2_label, export_rows)
        st.download_button(
            "PDF Report", data=pdf_buf,
            file_name=f"{company_name.replace(' ', '_')}_ratio_report.pdf",
            mime="application/pdf", use_container_width=True,
        )

    st.caption(
        "Reminder for students: these interpretations use general rules of thumb. "
        "Always compare a company's ratios against its own history and against "
        "industry peers, since 'good' and 'bad' vary a lot by sector."
    )
