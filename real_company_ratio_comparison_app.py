"""
Real-Company Ratio Comparison Tool — for student learning
------------------------------------------------------------
Pulls real financial statements for up to 4 public companies via
yfinance, computes the same 17 ratios across two fiscal years,
compares companies side-by-side, and benchmarks them against
illustrative industry-average ratios.

IMPORTANT — data quality note for students:
Yahoo Finance's data (via the free yfinance library) is convenient
but not always complete or perfectly labeled — some companies are
missing a line item, or Yahoo reports it under an unexpected name.
That's why every fetched number is shown in an EDITABLE table before
ratios are computed: always sanity-check the auto-filled figures
against the company's actual 10-K before drawing conclusions.

Run with:  streamlit run real_company_ratio_comparison_app.py
"""

import io

import pandas as pd
import streamlit as st
import yfinance as yf

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter as PAGE_SIZE
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

st.set_page_config(page_title="Real-Company Ratio Comparison Tool", page_icon="🏢", layout="wide")

# ============================================================
# GENERIC HELPERS (shared logic with the manual-entry tool)
# ============================================================

def safe_div(numerator, denominator):
    if numerator is None or denominator in (0, None):
        return None
    return numerator / denominator


def fmt(value, suffix=""):
    if value is None:
        return "N/A"
    if suffix == "%":
        return f"{value * 100:,.1f}%"
    if suffix == " days":
        return f"{value:,.0f} days"
    if suffix == "x":
        return f"{value:,.2f}x"
    return f"{value:,.2f}"


def fmt_delta(delta, suffix):
    if delta is None:
        return None
    if suffix == "%":
        return f"{delta * 100:+.1f} pp"
    if suffix == " days":
        return f"{delta:+.0f} days"
    if suffix == "x":
        return f"{delta:+.2f}x"
    return f"{delta:+.2f}"


NA_MSG = "N/A"

# ============================================================
# INTERPRETATION FUNCTIONS — same thresholds as the manual tool.
# Used here only to classify a company's value into a
# Strong / Moderate / Weak badge; the descriptive `guide` text
# on each RATIO_DEF is shown once per ratio instead of repeating
# full prose for every company.
# ============================================================

def _tier(v, good, ok):
    if v is None:
        return "na"
    if good(v):
        return "good"
    if ok(v):
        return "ok"
    return "warning"


def level_current_ratio(v):   return _tier(v, lambda x: x >= 2, lambda x: x >= 1)
def level_quick_ratio(v):     return _tier(v, lambda x: x >= 1, lambda x: x >= 0.7)
def level_cash_ratio(v):      return _tier(v, lambda x: x >= 0.5, lambda x: x >= 0.2)
def level_gross_margin(v):    return _tier(v, lambda x: x >= 0.40, lambda x: x >= 0.20)
def level_operating_margin(v):return _tier(v, lambda x: x >= 0.15, lambda x: x >= 0.05)
def level_net_margin(v):      return _tier(v, lambda x: x >= 0.10, lambda x: x >= 0.02)
def level_roa(v):             return _tier(v, lambda x: x >= 0.10, lambda x: x >= 0.03)
def level_roe(v):             return _tier(v, lambda x: x >= 0.15, lambda x: x >= 0.05)
def level_asset_turnover(v):  return _tier(v, lambda x: x >= 1.0, lambda x: x >= 0.5)
def level_inventory_turnover(v): return _tier(v, lambda x: x >= 8, lambda x: x >= 3)
def level_dio(v):              return "na" if v is None else ("good" if v <= 45 else ("ok" if v <= 90 else "warning"))
def level_receivables_turnover(v): return _tier(v, lambda x: x >= 10, lambda x: x >= 4)
def level_dso(v):               return "na" if v is None else ("good" if v <= 45 else ("ok" if v <= 90 else "warning"))
def level_debt_ratio(v):        return "na" if v is None else ("good" if v <= 0.4 else ("ok" if v <= 0.6 else "warning"))
def level_debt_to_equity(v):    return "na" if v is None else ("good" if v <= 1.0 else ("ok" if v <= 2.0 else "warning"))
def level_equity_multiplier(v): return "na" if v is None else ("good" if v <= 2.0 else ("ok" if v <= 3.0 else "warning"))
def level_interest_coverage(v): return _tier(v, lambda x: x >= 5, lambda x: x >= 1.5)


LEVEL_LABEL = {"good": "✅ Strong", "ok": "🟡 Moderate", "warning": "🔴 Weak", "na": "⚪ N/A"}
LEVEL_LABEL_PLAIN = {"good": "Strong", "ok": "Moderate", "warning": "Weak", "na": "N/A"}

# ============================================================
# RATIO DEFINITIONS (single source of truth)
# ============================================================
RATIO_DEFS = [
    dict(key="current_ratio", category="Liquidity", name="Current Ratio",
         formula="Current Assets ÷ Current Liabilities", suffix="", lower_better=False, level_fn=level_current_ratio,
         guide="Rule of thumb: ≥2 strong · 1–2 moderate · <1 weak (current liabilities exceed current assets)."),
    dict(key="quick_ratio", category="Liquidity", name="Quick Ratio (Acid-Test)",
         formula="(Current Assets − Inventory) ÷ Current Liabilities", suffix="", lower_better=False, level_fn=level_quick_ratio,
         guide="Rule of thumb: ≥1 strong · 0.7–1 moderate · <0.7 weak. Excludes inventory from liquid assets."),
    dict(key="cash_ratio", category="Liquidity", name="Cash Ratio",
         formula="Cash ÷ Current Liabilities", suffix="", lower_better=False, level_fn=level_cash_ratio,
         guide="Rule of thumb: ≥0.5 strong · 0.2–0.5 moderate · <0.2 weak. The most conservative liquidity measure."),

    dict(key="gross_margin", category="Profitability", name="Gross Profit Margin",
         formula="Gross Profit ÷ Sales", suffix="%", lower_better=False, level_fn=level_gross_margin,
         guide="Rule of thumb: ≥40% strong · 20–40% moderate · <20% weak. Varies enormously by industry."),
    dict(key="operating_margin", category="Profitability", name="Operating Profit Margin",
         formula="EBIT ÷ Sales", suffix="%", lower_better=False, level_fn=level_operating_margin,
         guide="Rule of thumb: ≥15% strong · 5–15% moderate · <5% weak."),
    dict(key="net_margin", category="Profitability", name="Net Profit Margin",
         formula="Net Income ÷ Sales", suffix="%", lower_better=False, level_fn=level_net_margin,
         guide="Rule of thumb: ≥10% strong · 2–10% moderate · <2% weak."),
    dict(key="roa", category="Profitability", name="Return on Assets (ROA)",
         formula="Net Income ÷ Total Assets", suffix="%", lower_better=False, level_fn=level_roa,
         guide="Rule of thumb: ≥10% strong · 3–10% moderate · <3% weak."),
    dict(key="roe", category="Profitability", name="Return on Equity (ROE)",
         formula="Net Income ÷ Shareholders' Equity", suffix="%", lower_better=False, level_fn=level_roe,
         guide="Rule of thumb: ≥15% strong · 5–15% moderate · <5% weak."),

    dict(key="asset_turnover", category="Efficiency", name="Asset Turnover",
         formula="Sales ÷ Total Assets", suffix="x", lower_better=False, level_fn=level_asset_turnover,
         guide="Rule of thumb: ≥1.0x strong · 0.5–1.0x moderate · <0.5x weak. Capital-intensive firms run lower."),
    dict(key="inventory_turnover", category="Efficiency", name="Inventory Turnover",
         formula="COGS ÷ Inventory", suffix="x", lower_better=False, level_fn=level_inventory_turnover,
         guide="Rule of thumb: ≥8x strong · 3–8x moderate · <3x weak."),
    dict(key="dio", category="Efficiency", name="Days Inventory Outstanding",
         formula="365 ÷ Inventory Turnover", suffix=" days", lower_better=True, level_fn=level_dio,
         guide="Rule of thumb: ≤45 days strong · 45–90 moderate · >90 weak (lower is better)."),
    dict(key="receivables_turnover", category="Efficiency", name="Receivables Turnover",
         formula="Sales ÷ Accounts Receivable", suffix="x", lower_better=False, level_fn=level_receivables_turnover,
         guide="Rule of thumb: ≥10x strong · 4–10x moderate · <4x weak."),
    dict(key="dso", category="Efficiency", name="Days Sales Outstanding (DSO)",
         formula="365 ÷ Receivables Turnover", suffix=" days", lower_better=True, level_fn=level_dso,
         guide="Rule of thumb: ≤45 days strong · 45–90 moderate · >90 weak (lower is better)."),

    dict(key="debt_ratio", category="Solvency", name="Debt Ratio",
         formula="Total Liabilities ÷ Total Assets", suffix="%", lower_better=True, level_fn=level_debt_ratio,
         guide="Rule of thumb: ≤40% strong · 40–60% moderate · >60% weak (lower is better)."),
    dict(key="debt_to_equity", category="Solvency", name="Debt-to-Equity Ratio",
         formula="Total Liabilities ÷ Equity", suffix="x", lower_better=True, level_fn=level_debt_to_equity,
         guide="Rule of thumb: ≤1.0x strong · 1–2x moderate · >2x weak (lower is better)."),
    dict(key="equity_multiplier", category="Solvency", name="Equity Multiplier",
         formula="Total Assets ÷ Equity", suffix="x", lower_better=True, level_fn=level_equity_multiplier,
         guide="Rule of thumb: ≤2.0x strong · 2–3x moderate · >3x weak (lower is better)."),
    dict(key="interest_coverage", category="Solvency", name="Interest Coverage Ratio",
         formula="EBIT ÷ Interest Expense", suffix="x", lower_better=False, level_fn=level_interest_coverage,
         guide="Rule of thumb: ≥5x strong · 1.5–5x moderate · <1.5x weak."),
]
CATEGORIES = ["Liquidity", "Profitability", "Efficiency", "Solvency"]
RATIO_BY_KEY = {r["key"]: r for r in RATIO_DEFS}

# ============================================================
# DERIVED FIGURES & RATIO CALCULATION
# (real-company data already has current_assets, total_liabilities,
#  gross_profit, ebit, net_income as reported line items — no need
#  to reconstruct them via a subtraction chain like the manual tool.)
# ============================================================

def compute_derived(v):
    v = dict(v)
    if v.get("current_assets") is None and all(v.get(k) is not None for k in ("cash", "receivables", "inventory")):
        v["current_assets"] = v["cash"] + v["receivables"] + v["inventory"]
    if v.get("total_liabilities") is None and v.get("total_assets") is not None and v.get("equity") is not None:
        v["total_liabilities"] = v["total_assets"] - v["equity"]
    if v.get("equity") is None and v.get("total_assets") is not None and v.get("total_liabilities") is not None:
        v["equity"] = v["total_assets"] - v["total_liabilities"]
    if v.get("gross_profit") is None and v.get("revenue") is not None and v.get("cogs") is not None:
        v["gross_profit"] = v["revenue"] - v["cogs"]
    return v


def compute_ratios(d):
    r = {}
    r["current_ratio"] = safe_div(d.get("current_assets"), d.get("current_liabilities"))
    r["quick_ratio"] = safe_div(
        None if d.get("current_assets") is None or d.get("inventory") is None else d["current_assets"] - d["inventory"],
        d.get("current_liabilities"),
    )
    r["cash_ratio"] = safe_div(d.get("cash"), d.get("current_liabilities"))

    r["gross_margin"] = safe_div(d.get("gross_profit"), d.get("revenue"))
    r["operating_margin"] = safe_div(d.get("ebit"), d.get("revenue"))
    r["net_margin"] = safe_div(d.get("net_income"), d.get("revenue"))
    r["roa"] = safe_div(d.get("net_income"), d.get("total_assets"))
    r["roe"] = safe_div(d.get("net_income"), d.get("equity"))

    r["asset_turnover"] = safe_div(d.get("revenue"), d.get("total_assets"))
    r["inventory_turnover"] = safe_div(d.get("cogs"), d.get("inventory"))
    r["dio"] = safe_div(365, r["inventory_turnover"]) if r["inventory_turnover"] else None
    r["receivables_turnover"] = safe_div(d.get("revenue"), d.get("receivables"))
    r["dso"] = safe_div(365, r["receivables_turnover"]) if r["receivables_turnover"] else None

    r["debt_ratio"] = safe_div(d.get("total_liabilities"), d.get("total_assets"))
    r["debt_to_equity"] = safe_div(d.get("total_liabilities"), d.get("equity"))
    r["equity_multiplier"] = safe_div(d.get("total_assets"), d.get("equity"))
    r["interest_coverage"] = safe_div(d.get("ebit"), d.get("interest_expense"))
    return r


# ============================================================
# INDUSTRY BENCHMARKS (illustrative reference points, editable in-app)
# These are general educational approximations, NOT official current
# industry data — there's no reliable free live source for a full
# ratio suite. For rigorous work, point students to Damodaran Online
# (NYU Stern, pages.stern.nyu.edu/~adamodar) or paid sources like
# CSIMarket / IBISWorld, and let them overwrite these numbers.
# ============================================================
BENCHMARKS = {
    "Technology": dict(current_ratio=2.5, quick_ratio=2.2, cash_ratio=1.0, gross_margin=0.55, operating_margin=0.20,
                        net_margin=0.15, roa=0.10, roe=0.20, asset_turnover=0.6, inventory_turnover=8, dio=45,
                        receivables_turnover=8, dso=45, debt_ratio=0.40, debt_to_equity=0.7, equity_multiplier=1.8,
                        interest_coverage=15),
    "Healthcare": dict(current_ratio=1.8, quick_ratio=1.4, cash_ratio=0.5, gross_margin=0.55, operating_margin=0.15,
                        net_margin=0.10, roa=0.07, roe=0.15, asset_turnover=0.7, inventory_turnover=5, dio=73,
                        receivables_turnover=7, dso=52, debt_ratio=0.45, debt_to_equity=0.8, equity_multiplier=1.9,
                        interest_coverage=8),
    "Financial Services": dict(current_ratio=None, quick_ratio=None, cash_ratio=None, gross_margin=None,
                                operating_margin=None, net_margin=0.20, roa=0.012, roe=0.11, asset_turnover=None,
                                inventory_turnover=None, dio=None, receivables_turnover=None, dso=None,
                                debt_ratio=0.88, debt_to_equity=7.0, equity_multiplier=10.0, interest_coverage=2.5),
    "Consumer Cyclical": dict(current_ratio=1.4, quick_ratio=0.5, cash_ratio=0.2, gross_margin=0.35, operating_margin=0.08,
                               net_margin=0.05, roa=0.06, roe=0.18, asset_turnover=1.8, inventory_turnover=6, dio=61,
                               receivables_turnover=20, dso=18, debt_ratio=0.55, debt_to_equity=1.3, equity_multiplier=2.5,
                               interest_coverage=6),
    "Consumer Defensive": dict(current_ratio=1.0, quick_ratio=0.5, cash_ratio=0.15, gross_margin=0.30, operating_margin=0.10,
                                net_margin=0.06, roa=0.07, roe=0.20, asset_turnover=1.2, inventory_turnover=8, dio=46,
                                receivables_turnover=12, dso=30, debt_ratio=0.55, debt_to_equity=1.4, equity_multiplier=2.6,
                                interest_coverage=8),
    "Industrials": dict(current_ratio=1.5, quick_ratio=1.0, cash_ratio=0.3, gross_margin=0.30, operating_margin=0.11,
                         net_margin=0.07, roa=0.06, roe=0.15, asset_turnover=0.8, inventory_turnover=6, dio=61,
                         receivables_turnover=8, dso=46, debt_ratio=0.50, debt_to_equity=1.0, equity_multiplier=2.2,
                         interest_coverage=7),
    "Energy": dict(current_ratio=1.2, quick_ratio=0.9, cash_ratio=0.3, gross_margin=0.35, operating_margin=0.15,
                    net_margin=0.08, roa=0.06, roe=0.12, asset_turnover=0.5, inventory_turnover=10, dio=37,
                    receivables_turnover=10, dso=37, debt_ratio=0.45, debt_to_equity=0.8, equity_multiplier=1.9,
                    interest_coverage=6),
    "Utilities": dict(current_ratio=0.9, quick_ratio=0.7, cash_ratio=0.1, gross_margin=0.40, operating_margin=0.20,
                       net_margin=0.10, roa=0.03, roe=0.10, asset_turnover=0.3, inventory_turnover=8, dio=46,
                       receivables_turnover=10, dso=37, debt_ratio=0.60, debt_to_equity=1.5, equity_multiplier=2.8,
                       interest_coverage=3.5),
    "Real Estate": dict(current_ratio=1.0, quick_ratio=0.8, cash_ratio=0.2, gross_margin=0.50, operating_margin=0.35,
                         net_margin=0.20, roa=0.03, roe=0.08, asset_turnover=0.15, inventory_turnover=None, dio=None,
                         receivables_turnover=15, dso=24, debt_ratio=0.55, debt_to_equity=1.3, equity_multiplier=2.3,
                         interest_coverage=3),
    "Basic Materials": dict(current_ratio=1.6, quick_ratio=1.0, cash_ratio=0.3, gross_margin=0.25, operating_margin=0.12,
                             net_margin=0.07, roa=0.06, roe=0.13, asset_turnover=0.7, inventory_turnover=6, dio=61,
                             receivables_turnover=8, dso=46, debt_ratio=0.45, debt_to_equity=0.8, equity_multiplier=1.9,
                             interest_coverage=6),
    "Communication Services": dict(current_ratio=1.0, quick_ratio=0.9, cash_ratio=0.3, gross_margin=0.55, operating_margin=0.18,
                                    net_margin=0.10, roa=0.05, roe=0.12, asset_turnover=0.4, inventory_turnover=None,
                                    dio=None, receivables_turnover=8, dso=46, debt_ratio=0.55, debt_to_equity=1.3,
                                    equity_multiplier=2.3, interest_coverage=5),
    "General / Unknown": dict(current_ratio=1.5, quick_ratio=1.0, cash_ratio=0.3, gross_margin=0.35, operating_margin=0.12,
                               net_margin=0.08, roa=0.06, roe=0.14, asset_turnover=0.8, inventory_turnover=7, dio=52,
                               receivables_turnover=9, dso=40, debt_ratio=0.50, debt_to_equity=1.0, equity_multiplier=2.0,
                               interest_coverage=6),
}

# ============================================================
# YFINANCE FETCH LOGIC
# ============================================================
BS_CANDIDATES = {
    "cash": ["Cash And Cash Equivalents", "Cash Cash Equivalents And Short Term Investments", "Cash"],
    "receivables": ["Receivables", "Accounts Receivable", "Net Receivables"],
    "inventory": ["Inventory"],
    "current_assets": ["Current Assets", "Total Current Assets"],
    "total_assets": ["Total Assets"],
    "current_liabilities": ["Current Liabilities", "Total Current Liabilities"],
    "total_liabilities": ["Total Liabilities Net Minority Interest", "Total Liab"],
    "equity": ["Common Stock Equity", "Stockholders Equity", "Total Equity Gross Minority Interest", "Total Stockholder Equity"],
}
IS_CANDIDATES = {
    "revenue": ["Total Revenue", "Operating Revenue"],
    "cogs": ["Cost Of Revenue", "Reconciled Cost Of Revenue"],
    "gross_profit": ["Gross Profit"],
    "ebit": ["Operating Income", "EBIT", "Total Operating Income As Reported"],
    "interest_expense": ["Interest Expense", "Interest Expense Non Operating"],
    "net_income": ["Net Income", "Net Income Common Stockholders", "Net Income Continuous Operations"],
}
FIELD_LABELS = {
    "cash": "Cash & Cash Equivalents", "receivables": "Accounts Receivable", "inventory": "Inventory",
    "current_assets": "Total Current Assets", "total_assets": "Total Assets",
    "current_liabilities": "Total Current Liabilities", "total_liabilities": "Total Liabilities", "equity": "Shareholders' Equity",
    "revenue": "Net Sales / Revenue", "cogs": "Cost of Goods Sold (COGS)", "gross_profit": "Gross Profit",
    "ebit": "Operating Income (EBIT)", "interest_expense": "Interest Expense", "net_income": "Net Income",
}
BS_FIELDS = ["cash", "receivables", "inventory", "current_assets", "total_assets", "current_liabilities",
             "total_liabilities", "equity"]
IS_FIELDS = ["revenue", "cogs", "gross_profit", "ebit", "interest_expense", "net_income"]


def _period_label(col):
    try:
        return f"FY{col.year}"
    except Exception:
        return str(col)


def _lookup(df, candidates, col, missing_list, field_key):
    if df is None or df.empty or col is None:
        missing_list.append(field_key)
        return None
    for cand in candidates:
        if cand in df.index:
            try:
                val = df.loc[cand, col]
            except Exception:
                continue
            if pd.notna(val):
                return float(val)
    missing_list.append(field_key)
    return None


@st.cache_data(ttl=3600, show_spinner=False)
def fetch_company(ticker_symbol):
    """Fetch two fiscal years of financials for one ticker. Never raises —
    returns a dict with an 'error' key set if the whole fetch failed."""
    ticker_symbol = ticker_symbol.strip().upper()
    try:
        t = yf.Ticker(ticker_symbol)
    except Exception as e:
        return dict(ticker=ticker_symbol, error=f"Could not create Ticker object: {e}")

    try:
        info = t.info or {}
    except Exception:
        info = {}
    name = info.get("longName") or info.get("shortName") or ticker_symbol
    sector = info.get("sector") or "Unknown"
    currency = info.get("currency") or "USD"

    try:
        bs = t.balance_sheet
    except Exception:
        bs = pd.DataFrame()
    try:
        inc = t.income_stmt
    except Exception:
        try:
            inc = t.financials
        except Exception:
            inc = pd.DataFrame()

    if bs is None:
        bs = pd.DataFrame()
    if inc is None:
        inc = pd.DataFrame()

    if bs.empty and inc.empty:
        return dict(ticker=ticker_symbol, error="No financial statement data returned by Yahoo Finance for this ticker.")

    bs_periods = sorted(list(bs.columns), reverse=True) if not bs.empty else []
    inc_periods = sorted(list(inc.columns), reverse=True) if not inc.empty else []

    def year_values(idx):
        vals, missing = {}, []
        bs_col = bs_periods[idx] if idx < len(bs_periods) else None
        inc_col = inc_periods[idx] if idx < len(inc_periods) else None
        for key in BS_FIELDS:
            vals[key] = _lookup(bs, BS_CANDIDATES[key], bs_col, missing, key)
        for key in IS_FIELDS:
            vals[key] = _lookup(inc, IS_CANDIDATES[key], inc_col, missing, key)
        if vals.get("interest_expense") is not None:
            vals["interest_expense"] = abs(vals["interest_expense"])
        return vals, missing

    values_y2, missing_y2 = year_values(0)
    values_y1, missing_y1 = year_values(1) if (len(bs_periods) > 1 or len(inc_periods) > 1) else ({k: None for k in BS_FIELDS + IS_FIELDS}, BS_FIELDS + IS_FIELDS)

    y2_label = _period_label(inc_periods[0]) if inc_periods else (_period_label(bs_periods[0]) if bs_periods else "Year 2")
    y1_label = _period_label(inc_periods[1]) if len(inc_periods) > 1 else (_period_label(bs_periods[1]) if len(bs_periods) > 1 else "Year 1")

    return dict(
        ticker=ticker_symbol, name=name, sector=sector, currency=currency,
        values_y1=values_y1, values_y2=values_y2,
        y1_label=y1_label, y2_label=y2_label,
        missing_y1=missing_y1, missing_y2=missing_y2,
        error=None,
    )


# ============================================================
# EXCEL EXPORT
# ============================================================

def build_excel_report(companies, benchmark_sector, benchmark, matrix_rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Ratio Comparison"

    title_font = Font(name="Arial", size=14, bold=True)
    subtitle_font = Font(name="Arial", size=10, italic=True, color="666666")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    normal_font = Font(name="Arial", size=10)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    n_cols = 2 + 2 * len(companies) + 1
    last_col_letter = get_column_letter(n_cols)
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = "Real-Company Ratio Comparison"
    ws["A1"].font = title_font
    ws.merge_cells(f"A2:{last_col_letter}2")
    ws["A2"] = f"Companies: {', '.join(c['ticker'] for c in companies)}  |  Benchmark sector: {benchmark_sector}"
    ws["A2"].font = subtitle_font

    headers = ["Category", "Ratio"]
    for c in companies:
        headers += [f"{c['ticker']} {c['y1_label']}", f"{c['ticker']} {c['y2_label']}"]
    headers += ["Industry Benchmark"]

    header_row = 4
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    number_formats = {"": "0.00", "%": "0.0%", "x": '0.00"x"', " days": '0" days"'}
    r = header_row
    for row in matrix_rows:
        r += 1
        nf = number_formats.get(row["suffix"], "0.00")
        ws.cell(row=r, column=1, value=row["category"]).font = normal_font
        ws.cell(row=r, column=2, value=row["ratio"]).font = normal_font
        col_idx = 3
        for val in row["company_values"]:
            cell = ws.cell(row=r, column=col_idx, value=("N/A" if val is None else val))
            cell.font = normal_font
            if val is not None:
                cell.number_format = nf
            cell.alignment = Alignment(horizontal="center")
            col_idx += 1
        bcell = ws.cell(row=r, column=col_idx, value=("N/A" if row["benchmark"] is None else row["benchmark"]))
        bcell.font = normal_font
        if row["benchmark"] is not None:
            bcell.number_format = nf
        bcell.alignment = Alignment(horizontal="center")
        for ci in range(1, n_cols + 1):
            ws.cell(row=r, column=ci).border = border

    widths = [14, 28] + [14] * (2 * len(companies)) + [18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "C5"

    # Raw data, one sheet per company
    for c in companies:
        sheet_name = c["ticker"][:31]
        ws2 = wb.create_sheet(sheet_name)
        ws2.merge_cells("A1:C1")
        ws2["A1"] = f"{c['name']} ({c['ticker']}) — Raw Inputs"
        ws2["A1"].font = title_font
        raw_headers = ["Line Item", c["y1_label"], c["y2_label"]]
        for ci, h in enumerate(raw_headers, start=1):
            cell = ws2.cell(row=3, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        rr = 3
        for key in BS_FIELDS + IS_FIELDS:
            rr += 1
            ws2.cell(row=rr, column=1, value=FIELD_LABELS[key]).font = normal_font
            for ci, yv in ((2, c["values_y1"].get(key)), (3, c["values_y2"].get(key))):
                cell = ws2.cell(row=rr, column=ci, value=("N/A" if yv is None else yv))
                cell.font = normal_font
                if yv is not None:
                    cell.number_format = "#,##0"
                cell.border = border
            ws2.cell(row=rr, column=1).border = border
        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 16
        ws2.column_dimensions["C"].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ============================================================
# PDF EXPORT
# ============================================================

def build_pdf_report(companies, benchmark_sector, matrix_rows):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(PAGE_SIZE),
        leftMargin=30, rightMargin=30, topMargin=36, bottomMargin=30,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("ReportTitle", parent=styles["Title"], fontSize=16, spaceAfter=4)
    sub_style = ParagraphStyle("ReportSub", parent=styles["Normal"], fontSize=9, textColor=colors.grey, spaceAfter=14)
    section_style = ParagraphStyle("Section", parent=styles["Heading2"], fontSize=12, spaceBefore=14, spaceAfter=6,
                                    textColor=colors.HexColor("#1F4E78"))
    footnote_style = ParagraphStyle("Footnote", parent=styles["Normal"], fontSize=7.5, textColor=colors.grey)

    tickers = ", ".join(c["ticker"] for c in companies)
    story = [
        Paragraph("Real-Company Ratio Comparison", title_style),
        Paragraph(f"Companies: {tickers}  |  Benchmark sector: {benchmark_sector}", sub_style),
    ]

    col_widths = [140] + [55] * (2 * len(companies)) + [60]
    for cat in CATEGORIES:
        rows = [r for r in matrix_rows if r["category"] == cat]
        if not rows:
            continue
        story.append(Paragraph(cat, section_style))
        header = ["Ratio"]
        for c in companies:
            header += [c["y1_label"], c["y2_label"]]
        header += ["Bench."]
        ticker_header = [""]
        for c in companies:
            ticker_header += [c["ticker"], ""]
        ticker_header += [""]
        data = [ticker_header, header]
        for r in rows:
            line = [r["ratio"]]
            for val in r["company_values"]:
                line.append(fmt(val, r["suffix"]))
            line.append(fmt(r["benchmark"], r["suffix"]))
            data.append(line)
        table = Table(data, colWidths=col_widths[: len(header)], repeatRows=2)
        style_cmds = [
            ("SPAN", (0, 0), (0, 1)),
            ("BACKGROUND", (0, 0), (-1, 1), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 1), colors.white),
            ("FONTNAME", (0, 0), (-1, 1), "Helvetica-Bold"),
            ("FONTNAME", (0, 2), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
            ("ROWBACKGROUNDS", (0, 2), (-1, -1), [colors.white, colors.HexColor("#F2F6FA")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]
        col = 1
        for _ in companies:
            style_cmds.append(("SPAN", (col, 0), (col + 1, 0)))
            col += 2
        table.setStyle(TableStyle(style_cmds))
        story.append(table)

    story.append(Spacer(1, 12))
    story.append(Paragraph(
        "Industry benchmark values are general educational reference points, not live data. "
        "Assessment thresholds are classroom rules of thumb; real-world norms vary by industry. "
        "Generated by the Real-Company Ratio Comparison Tool.",
        footnote_style,
    ))
    doc.build(story)
    buf.seek(0)
    return buf


# ============================================================
# SIDEBAR
# ============================================================
with st.sidebar:
    st.header("📘 How to use this tool")
    st.markdown(
        """
        1. Enter up to **4 stock tickers** (comma-separated) and click
           **Fetch Data**. Figures come from Yahoo Finance via the
           free `yfinance` library.
        2. **Always review the editable tables** that appear after
           fetching — free data sources sometimes miss a line item
           or label it unexpectedly. Fix anything that looks wrong
           before trusting the ratios.
        3. Pick (or keep the auto-detected) **industry benchmark**
           and adjust its values if you have more precise figures
           (e.g. from Damodaran Online at NYU Stern).
        4. Browse the ratio tabs to compare companies against each
           other and against the benchmark, then download the
           **Summary Dashboard** as CSV, Excel, or PDF.

        *Data quality note:* yfinance is free and convenient but not
        always complete — this is a real limitation of free data, not
        a bug. Cross-check anything surprising against the company's
        actual 10-K/annual report.
        """
    )
    st.caption("Ratio thresholds are general classroom rules of thumb; benchmark values are illustrative, editable estimates — not official industry data.")

st.title("🏢 Real-Company Ratio Comparison Tool")
st.caption("Pull real financial statements, compare companies over two years, and benchmark against industry norms.")

# ============================================================
# TICKER INPUT & FETCH
# ============================================================
default_tickers = "AAPL, MSFT"
tickers_input = st.text_input("Stock tickers (comma-separated, up to 4)", value=default_tickers)
fetch_clicked = st.button("🔄 Fetch Data", type="primary")

tickers = [t.strip().upper() for t in tickers_input.split(",") if t.strip()][:4]

if fetch_clicked or "fetched_companies" not in st.session_state:
    if tickers:
        with st.spinner("Fetching financial statements from Yahoo Finance..."):
            fetched = [fetch_company(t) for t in tickers]
        st.session_state["fetched_companies"] = fetched
        st.session_state["fetched_tickers"] = tickers

companies_raw = st.session_state.get("fetched_companies", [])

errors = [c for c in companies_raw if c.get("error")]
companies_raw = [c for c in companies_raw if not c.get("error")]

for e in errors:
    st.error(f"**{e['ticker']}**: {e['error']}")

if not companies_raw:
    st.info("Enter one or more tickers above and click **Fetch Data** to get started.")
    st.stop()

# ============================================================
# REVIEW & EDIT FETCHED DATA
# ============================================================
st.subheader("📝 Review & Edit Fetched Data")
st.write(
    "Auto-fetched figures are pre-filled below. **Check them against the company's real filings** — "
    "any field yfinance couldn't find is set to 0 and flagged so you know to fill it in by hand."
)

companies = []
for c in companies_raw:
    with st.expander(f"{c['name']} ({c['ticker']}) — Sector: {c['sector']}", expanded=(len(companies_raw) <= 2)):
        col_bs, col_is = st.columns(2)

        with col_bs:
            st.markdown("**Balance Sheet**")
            bs_df = pd.DataFrame({
                "Item": [FIELD_LABELS[k] for k in BS_FIELDS],
                c["y1_label"]: [c["values_y1"].get(k) if c["values_y1"].get(k) is not None else 0.0 for k in BS_FIELDS],
                c["y2_label"]: [c["values_y2"].get(k) if c["values_y2"].get(k) is not None else 0.0 for k in BS_FIELDS],
            })
            edited_bs = st.data_editor(
                bs_df, key=f"bs_editor_{c['ticker']}", hide_index=True, num_rows="fixed", use_container_width=True,
                column_config={
                    "Item": st.column_config.TextColumn("Line Item", disabled=True),
                    c["y1_label"]: st.column_config.NumberColumn(c["y1_label"], format="%.0f"),
                    c["y2_label"]: st.column_config.NumberColumn(c["y2_label"], format="%.0f"),
                },
            )

        with col_is:
            st.markdown("**Income Statement**")
            is_df = pd.DataFrame({
                "Item": [FIELD_LABELS[k] for k in IS_FIELDS],
                c["y1_label"]: [c["values_y1"].get(k) if c["values_y1"].get(k) is not None else 0.0 for k in IS_FIELDS],
                c["y2_label"]: [c["values_y2"].get(k) if c["values_y2"].get(k) is not None else 0.0 for k in IS_FIELDS],
            })
            edited_is = st.data_editor(
                is_df, key=f"is_editor_{c['ticker']}", hide_index=True, num_rows="fixed", use_container_width=True,
                column_config={
                    "Item": st.column_config.TextColumn("Line Item", disabled=True),
                    c["y1_label"]: st.column_config.NumberColumn(c["y1_label"], format="%.0f"),
                    c["y2_label"]: st.column_config.NumberColumn(c["y2_label"], format="%.0f"),
                },
            )

        missing_msgs = []
        if c["missing_y1"]:
            missing_msgs.append(f"{c['y1_label']}: " + ", ".join(FIELD_LABELS[k] for k in c["missing_y1"]))
        if c["missing_y2"]:
            missing_msgs.append(f"{c['y2_label']}: " + ", ".join(FIELD_LABELS[k] for k in c["missing_y2"]))
        if missing_msgs:
            st.warning("⚠️ Not found automatically (defaulted to 0 — please fill in manually if known):\n\n" + "\n\n".join(missing_msgs))
        else:
            st.success("✅ All fields found automatically.")

        values_y1 = {k: float(v) for k, v in zip(BS_FIELDS, edited_bs[c["y1_label"]])}
        values_y1.update({k: float(v) for k, v in zip(IS_FIELDS, edited_is[c["y1_label"]])})
        values_y2 = {k: float(v) for k, v in zip(BS_FIELDS, edited_bs[c["y2_label"]])}
        values_y2.update({k: float(v) for k, v in zip(IS_FIELDS, edited_is[c["y2_label"]])})

        companies.append(dict(
            ticker=c["ticker"], name=c["name"], sector=c["sector"], currency=c["currency"],
            y1_label=c["y1_label"], y2_label=c["y2_label"],
            values_y1=values_y1, values_y2=values_y2,
        ))

for c in companies:
    d1 = compute_derived(c["values_y1"])
    d2 = compute_derived(c["values_y2"])
    c["ratios_y1"] = compute_ratios(d1)
    c["ratios_y2"] = compute_ratios(d2)

# ============================================================
# INDUSTRY BENCHMARK SELECTION
# ============================================================
st.subheader("📊 Industry Benchmark")
sector_options = list(BENCHMARKS.keys())
detected_sector = companies[0]["sector"] if companies[0]["sector"] in BENCHMARKS else "General / Unknown"
default_idx = sector_options.index(detected_sector) if detected_sector in sector_options else len(sector_options) - 1
benchmark_sector = st.selectbox(
    f"Benchmark sector (auto-detected from {companies[0]['ticker']}: {companies[0]['sector']})",
    sector_options, index=default_idx,
)
st.caption(
    "These are illustrative educational reference points, not official current industry data — "
    "edit any value below if you have more precise figures (e.g. from Damodaran Online at NYU Stern, "
    "CSIMarket, or IBISWorld)."
)

bench_defaults = BENCHMARKS[benchmark_sector]
bench_df = pd.DataFrame({
    "Ratio": [r["name"] for r in RATIO_DEFS],
    "Format": [{"": "plain", "%": "percent (enter as fraction, e.g. 0.35)", "x": "multiple",
                " days": "days"}[r["suffix"]] for r in RATIO_DEFS],
    "Benchmark Value": [bench_defaults.get(r["key"]) for r in RATIO_DEFS],
})
edited_bench = st.data_editor(
    bench_df, key=f"bench_editor_{benchmark_sector}", hide_index=True, num_rows="fixed", use_container_width=True,
    column_config={
        "Ratio": st.column_config.TextColumn(disabled=True),
        "Format": st.column_config.TextColumn(disabled=True),
        "Benchmark Value": st.column_config.NumberColumn(format="%.4f"),
    },
)
benchmark = {r["key"]: (float(v) if pd.notna(v) else None) for r, v in zip(RATIO_DEFS, edited_bench["Benchmark Value"])}

# ============================================================
# RENDERING HELPERS
# ============================================================

def render_ratio_comparison(rdef):
    st.markdown(f"#### {rdef['name']}")
    st.caption(f"Formula: {rdef['formula']}")
    st.caption(rdef["guide"])

    bench_val = benchmark.get(rdef["key"])
    rows = []
    chart_data = {}
    for c in companies:
        v1 = c["ratios_y1"][rdef["key"]]
        v2 = c["ratios_y2"][rdef["key"]]
        delta = (v2 - v1) if (v1 is not None and v2 is not None) else None
        vs_bench = (v2 - bench_val) if (v2 is not None and bench_val is not None) else None
        level = rdef["level_fn"](v2)
        rows.append({
            "Company": f"{c['ticker']}", c["y1_label"]: fmt(v1, rdef["suffix"]), c["y2_label"]: fmt(v2, rdef["suffix"]),
            "YoY Change": fmt_delta(delta, rdef["suffix"]) or "N/A",
            "vs Benchmark": fmt_delta(vs_bench, rdef["suffix"]) or "N/A",
            "Assessment": LEVEL_LABEL[level],
        })
        if v2 is not None:
            chart_data[c["ticker"]] = v2
    if bench_val is not None:
        chart_data["Benchmark"] = bench_val

    df = pd.DataFrame(rows)
    st.dataframe(df, use_container_width=True, hide_index=True)

    if chart_data:
        chart_df = pd.DataFrame({"Latest Value": chart_data})
        st.bar_chart(chart_df)
    st.write("")


def render_category_tab(category):
    defs = [r for r in RATIO_DEFS if r["category"] == category]
    for rdef in defs:
        render_ratio_comparison(rdef)
        st.divider()


# ============================================================
# TABS
# ============================================================
tab_liq, tab_prof, tab_eff, tab_solv, tab_summary = st.tabs(
    ["💧 Liquidity", "💰 Profitability", "⚙️ Efficiency", "🏦 Solvency", "📋 Summary Dashboard"]
)

with tab_liq:
    st.write("Liquidity ratios measure whether a company can meet its short-term obligations.")
    render_category_tab("Liquidity")

with tab_prof:
    st.write("Profitability ratios measure how well a company converts sales and assets into profit.")
    render_category_tab("Profitability")

with tab_eff:
    st.write("Efficiency ratios measure how well a company uses its assets to generate sales.")
    render_category_tab("Efficiency")

with tab_solv:
    st.write("Solvency ratios measure long-term financial risk and reliance on debt financing.")
    render_category_tab("Solvency")

with tab_summary:
    st.subheader("📋 Summary Dashboard")
    st.caption(f"Companies: {', '.join(c['ticker'] for c in companies)}  |  Benchmark: {benchmark_sector}")

    matrix_rows = []
    for rdef in RATIO_DEFS:
        company_values = []
        for c in companies:
            company_values.append(c["ratios_y1"][rdef["key"]])
            company_values.append(c["ratios_y2"][rdef["key"]])
        matrix_rows.append(dict(
            category=rdef["category"], ratio=rdef["name"], suffix=rdef["suffix"],
            company_values=company_values, benchmark=benchmark.get(rdef["key"]),
        ))

    display_cols = {"Category": [], "Ratio": []}
    for c in companies:
        display_cols[f"{c['ticker']} {c['y1_label']}"] = []
        display_cols[f"{c['ticker']} {c['y2_label']}"] = []
    display_cols["Industry Benchmark"] = []

    for row in matrix_rows:
        display_cols["Category"].append(row["category"])
        display_cols["Ratio"].append(row["ratio"])
        ci = 0
        for c in companies:
            display_cols[f"{c['ticker']} {c['y1_label']}"].append(fmt(row["company_values"][ci], row["suffix"]))
            display_cols[f"{c['ticker']} {c['y2_label']}"].append(fmt(row["company_values"][ci + 1], row["suffix"]))
            ci += 2
        display_cols["Industry Benchmark"].append(fmt(row["benchmark"], row["suffix"]))

    df_display = pd.DataFrame(display_cols)
    st.dataframe(df_display, use_container_width=True, hide_index=True)

    st.markdown("#### ⬇️ Download this comparison")
    dl1, dl2, dl3 = st.columns(3)
    with dl1:
        csv_bytes = df_display.to_csv(index=False).encode("utf-8")
        st.download_button("CSV", data=csv_bytes, file_name="company_ratio_comparison.csv",
                            mime="text/csv", use_container_width=True)
    with dl2:
        excel_buf = build_excel_report(companies, benchmark_sector, benchmark, matrix_rows)
        st.download_button("Excel (.xlsx)", data=excel_buf, file_name="company_ratio_comparison.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True)
    with dl3:
        pdf_buf = build_pdf_report(companies, benchmark_sector, matrix_rows)
        st.download_button("PDF Report", data=pdf_buf, file_name="company_ratio_comparison.pdf",
                            mime="application/pdf", use_container_width=True)

    st.caption(
        "Reminder: industry benchmark values are illustrative educational reference points, not live "
        "data, and interpretation thresholds are general classroom rules of thumb. Always cross-check "
        "unusual figures against the company's actual financial filings."
    )
